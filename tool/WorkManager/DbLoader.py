from DbMgr import DbMgr
from WorkRec import WorkRec
import openpyxl
import os


class DbLoader:
    def __init__(self):
        self.dbm = DbMgr()
        self.target_dir = 'D:\\svn\\py_wsp\\tool\\WorkManager'
        self.work_rec_file_name = 'workrec.db.txt'
        self.work_rec_file_path = self.target_dir + '\\' + self.work_rec_file_name
        self.excel_name = '工作记录表.xlsx'
        self.excel_path = self.target_dir + '\\' + self.excel_name
        self.f = None

    def load(self):
        # 打开格式化工作记录文件
        dbm = self.dbm
        dbm.create_workrec_table()
        f = open(self.work_rec_file_path, 'r', encoding='utf-8')
        for line in f:
            work_rec = WorkRec().load_from_str(line)
            dbm.insert(work_rec)
        f.close()

    def to_excel(self):
        if os.path.exists(self.excel_path):
            os.remove(self.excel_path)
        excel = openpyxl.Workbook()
        sheet = excel.create_sheet('双周报-2023-01-01~2023-01-14', 0)
        title_list = ['项目名称', '客户名', '工作事项或问题', '统计耗时（h）']
        line_base = 1
        row_base = 1
        line = line_base
        row = row_base
        for title in title_list:
            sheet.cell(line, row, title)
            row += 1

        dbm = self.dbm
        recs = dbm.select_double_week_report()
        for rec in recs:
            row = row_base
            line += 1
            for v in rec:
                sheet.cell(line, row, v)
                row += 1
        excel.save(self.excel_path)
        excel.close()


if __name__ == '__main__':
    dbl = DbLoader()
    dbl.load()
    dbl.to_excel()
