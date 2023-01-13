import openpyxl

if __name__ == '__main__':
    excel = openpyxl.Workbook()
    sheet = excel.create_sheet('demo_sheet',0)
    cell11 = sheet.cell(row=1, column=1)

    # sheet.row_dimensions[1].height = 22
    # sheet.column_dimensions['A'].width = 25
    cell11.value = "demo table title"
    # 居中
    cell11.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    excel.save('工作记录测试.xlsx')
